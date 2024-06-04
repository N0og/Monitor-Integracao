import os
from consts import LOGS_PATH
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment, numbers
from consts import VERSAO_ATUAL, VERSAO_MONITOR_ATUAL
from openpyxl.utils import get_column_letter
from datetime import date, datetime

def check_path_files(add, m):
        arquivos = os.listdir(add)
        for arquivo in arquivos:
            if arquivo == f"LOTES-{m}-ATEND-11-01-2024.xlsx":
                os.remove(os.path.join(add, arquivo))

#Checagem para sobrescrição de log.
def check_log():
    for log in os.listdir(LOGS_PATH):
        if date.today().strftime("%d-%m-%Y") in log:
            with open(os.path.join(LOGS_PATH, log), "w") as log_file:
                log_file.write("")
                log_file.close()

def diferenca_dias(dt1, dt2):
    if dt1 != 'SEM REDE' and dt2 != 'SEM REDE':
        formato = "%d/%m/%Y"
        
        if type(dt1) == str:
            dt1 = datetime.strptime(dt1, formato)
        if type(dt2) == str: 
            dt2 = datetime.strptime(dt2, formato)
        
        diferenca = abs((dt2 - dt1).days)
        
        return diferenca
    else:
        return 0
    
def header(
            aba,
            ultima_col,
            titulo = "MONITORAMENTO DE INTEGRAÇÃO | PEC x ATENDSAÚDE",
            sub_titulo = f"Versão mais recente do AtendSaúde: {VERSAO_ATUAL}   -   EXTRAÇÃO REALIZADA: {datetime.today().strftime('%d-%m-%Y')}     -   VIA: MONITORAMENTO-APOIO."
            ):
        # HEADER
    aba["A1"].value = titulo
    aba["A1"].alignment = Alignment(horizontal='center', vertical='center')
    aba["A1"].font = Font(name='Bahnschrift', size=24, bold=True, color="FFFFFF")
    aba["A1"].fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid")

    # HEADER
    aba["A4"].value = sub_titulo
    aba["A4"].alignment = Alignment(horizontal='center', vertical='center')
    aba["A4"].font = Font(name='Bahnschrift', size=10, bold=True, italic=True, color="FFFFFF")
    aba["A4"].fill = PatternFill(
        start_color="363840", end_color="363840", fill_type="solid")

    # HEADER
    aba.merge_cells(f"A1:{ultima_col}3")
    aba.merge_cells(f"A4:{ultima_col}4")




def dateStyle(aba, coluna, ultima_linha=None, ultima_coluna=None):

    last_col = ultima_coluna if ultima_coluna else aba.max_column
    last_row = ultima_linha if ultima_linha else aba.max_row


    for row in aba.iter_rows(min_row=6, max_row=last_row, min_col=coluna, max_col=coluna):
        for cell in row:
          cell.number_format = numbers.FORMAT_DATE_DDMMYY

def finish(aba, ultima_linha=None, ultima_coluna=None):

    last_col = ultima_coluna if ultima_coluna else aba.max_column
    last_row = ultima_linha if ultima_linha else aba.max_row

    # Centraliza
    for row in aba.iter_rows(min_row=6, max_row=last_row, min_col=1, max_col=last_col):
        for cell in row:
            cell.alignment = Alignment(
                horizontal='center', vertical='center')

    
    # Torna branca toda célula vázia.
    fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    for linha in range(1, last_row + 100):
        for coluna in range(1, last_row + 100):
            celula = aba.cell(row=linha, column=coluna)
            if celula.value is None:
                celula.fill = fill

    # Definir borda da planilha.
    border_style = Side(border_style='thin', color='000000')
    border = Border(top=border_style, bottom=border_style,
                    left=border_style, right=border_style)
    cell_select = aba[f'A5:{get_column_letter(last_col)}{last_row}']

    for linha in cell_select:
        for celula in linha:
            celula.border = border

    # Leg.
    aba[f"A{last_row+3}"].value = f"INTEGRAÇÃO|Monitoramento por APOIO INSTITUCIONAL - NOVETECH. v{VERSAO_MONITOR_ATUAL}"
    aba[f"A{last_row+3}"].font = Font(
        name='Calibri', size=9, bold=True, italic=True)