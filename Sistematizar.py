import pandas as pd
from consts import EXTRACT_PATH, OUTPUT_PATH, EAS_IPS
from os import path, listdir
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from util import diferenca_dias, header, finish, dateStyle

DATA_ATUAL = datetime.today().strftime("%d-%m-%Y")

class Lote:
    def __init__(self, uf:str, municipio:str, unidade:str, nu_lote:str, 
                 dt_lote:str, comp:str, inval:str, orig:str, hr_lote:str = None, env_por:str=None, valid:str=None, tipo:str=None, unidade_ext:str=None) -> None:
        self._no_uf = uf.upper()
        self._no_municipio = municipio.upper()
        self._unidade_ext = unidade_ext
        self._no_unidade = unidade
        self._nu_lote = nu_lote
        self._data = dt_lote if dt_lote == 'SEM REDE' or type(dt_lote) != str else datetime.strptime(dt_lote, '%d/%m/%Y') 
        self._hora = hr_lote
        self._comp = comp
        self._env_por = env_por
        self._valid = valid
        self._invalid = inval
        self._tipo = tipo
        self._origem = orig
    
    def getDadosPec(self):
        return [self._no_uf, self._no_municipio, self._unidade_ext, self._no_unidade, self._nu_lote, self._data, self._hora, self._comp, self._env_por, self._valid, self._invalid, self._tipo]
    
    def getDadosAtend(self):
        return [self._no_uf, self._no_municipio, self._no_unidade, self._nu_lote, self._data, self._comp, self._valid, self._invalid]

class Sistematizar:
    def __init__(self) -> None:
        self.lotes ={   "PEC": dict(), 
                        "ATEND": dict()}
        
        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)
        self.aba_analitico_clientes = self.workbook.create_sheet("ANALÍTICO CLIENTES AUTOMAÇÃO") 
        self.aba_detalhado_pec = self.workbook.create_sheet("DETALHADO PEC")
        self.aba_detalhado_atend = self.workbook.create_sheet("DETALHADO ATEND")

        self.workbook.active = self.aba_analitico_clientes
        
        #Carregar os lotes disponíveis em cada município.
        for self.uf, self.municipios in EAS_IPS.items():
            for self.municipio in self.municipios.keys():
                name_file_pec = f"LOTES-{self.municipio}-PEC-{DATA_ATUAL}.xlsx"
                name_file_atend = f"LOTES-{self.municipio}-ATEND-{DATA_ATUAL}.xlsx"        
                municipio_path = path.join(EXTRACT_PATH, self.uf, self.municipio)
                if not path.isdir(municipio_path):
                        continue


                for file in listdir(municipio_path):
                        if name_file_pec == file :
                            self.open_xlsx(pd.read_excel(path.join(municipio_path, name_file_pec)), "PEC")
                        if name_file_atend == file:
                            self.open_xlsx(pd.read_excel(path.join(municipio_path, name_file_atend)), "ATEND")

        self.sistematizar_consolidado()
        self.sistematizar_analitico_pec()
        self.sistematizar_analitico_atend()

        self.workbook.save(path.join(OUTPUT_PATH, f"MONITORAMENTOLOTES-{DATA_ATUAL}.xlsx"))
                                 
                                    
    def open_xlsx(self, xlsx:pd.DataFrame, orig:str):

        for ind, row in xlsx.iterrows():
            try:
                self.lotes[orig][self.municipio]
            except KeyError:
                self.lotes[orig][self.municipio] = dict()
            
            try:
                self.lotes[orig][self.municipio][row["UNIDADE"]]
            except KeyError:
                self.lotes[orig][self.municipio][row["UNIDADE"]] = []

            if orig == "ATEND":
                self.lotes[orig][self.municipio][row["UNIDADE"]].append(Lote(self.uf, self.municipio, row["UNIDADE"], row["LOTE"], row["DATA"], row["PERIODO"], row["INVALIDAS"], orig, valid=row['VALIDAS']))
            elif orig == "PEC":
                self.lotes[orig][self.municipio][row["UNIDADE"]].append(Lote(self.uf, self.municipio, row["UNIDADE"], row["LOTE"], row["DATA"], row["PERIODO"], row["INVALIDAS"], orig, row['HORA'], row["ENVIADO POR"], row['VALIDAS'], row['TIPO'], row["UNIDADE EXTRACAO"]))
            else:
                pass
    

    def sistematizar_consolidado(self):

        COLUNAS_CONSOLIDADO = ["UF", "MUNICIPIO", "UNIDADE", "ULTIMO LOTE GERADO", "ULTIMO LOTE IMPORTADO", "DATA DO U. LOTE G.", "DATA DO U. LOTE I.", "DIAS SEM IMPORTAR", "COMPETÊNCIA", "FICHAS INVÁLIDAS PEC", "FICHAS INVÁLIDAS ATEND", "STATUS"]

        header(self.aba_analitico_clientes, get_column_letter(len(COLUNAS_CONSOLIDADO)))

        for col, valor in enumerate(COLUNAS_CONSOLIDADO):
            celula = self.aba_analitico_clientes.cell(row=5, column=col+1, value=valor)
            celula.alignment = Alignment(horizontal='center', vertical='center')
            celula.font = Font(name='Calibri', size=11, bold=True, color="000000")
            celula.fill = PatternFill(start_color="c0c0c0", end_color="c0c0c0", fill_type="solid")
    
        row = 5
        
        for municipio_atend, unidades_atend in self.lotes["ATEND"].items():
            for unidade_atend, lotes_atend in unidades_atend.items():
                flag_lote_incidencia = False
                for lote_atend in lotes_atend:
                    row += 1
                    flag_lote = False
                    flag_data = False
                    flag_inv_pec = True
                    flag_inv_atend = True

                    conteudo = [lote_atend._no_uf, lote_atend._no_municipio, lote_atend._no_unidade, 
                                lote_atend._nu_lote, "-", lote_atend._data, "-", "-",
                                lote_atend._comp, "-", lote_atend._invalid, "-"]
                    
                    for municipio_pec, unidades_pec in self.lotes["PEC"].items():
                        if municipio_atend == municipio_pec:
                            for unidade_pec, lotes_pec in unidades_pec.items():
                                if unidade_pec == unidade_atend:
                                    datas_convertidas = []
                                    for lote_pec in lotes_pec:
                                       datas_convertidas.append(lote_pec._data)
                                    try:
                                        data_mais_recente = max(datas_convertidas)
                                        ultimos = [(lote_pec._nu_lote, lote_pec._data) for lote_pec in lotes_pec if lote_pec._data == data_mais_recente]
                                        conteudo[4] = ultimos[0][0]
                                        conteudo[6] = ultimos[0][1]
                                        conteudo[7] = diferenca_dias(lote_atend._data, ultimos[0][1])

                                        if lote_atend._nu_lote == ultimos[0][0]:
                                            flag_lote = True   
                                            flag_lote_incidencia = True
                                            if diferenca_dias(lote_atend._data, ultimos[0][1]) < 2:
                                                conteudo[11] = "OK"
                                                flag_data = True       
                                                conteudo[9] = lote_pec._invalid
                                                if int(lote_pec._invalid) > 0:
                                                    flag_inv_pec = False
                                            else:
                                                conteudo[11] = "DATA DE INSERÇÃO DESATUALIZADA (DIAS SEM ATUALIZAR)"
                                                conteudo[9] = "-"
                                            break
                                    except ValueError:
                                        pass
                                            
                                         
                    try:
                        if int(lote_atend._invalid) > 0:
                            flag_inv_atend = False            
                    except ValueError as e:
                        pass      
                    
                    if not flag_lote_incidencia:
                        conteudo[11] = "LOTE FORA DA COMPETÊNCIA (1 OU MAIS MESES DESATUALIZADO)"
                        conteudo[9] = "-"

                    if conteudo[4] == "-":
                        conteudo[11] = "SEM CONEXÃO NO MOMENTO DA EXTRAÇÃO DESTE RELATÓRIO"

                    self.aba_analitico_clientes.append(conteudo)
                    self.aba_analitico_clientes.cell(row=row, column=12).fill = PatternFill(start_color='b6d2d9', end_color='b6d2d9', fill_type='solid')
                    if flag_lote:
                        self.aba_analitico_clientes.cell(row=row, column=4).fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
                        self.aba_analitico_clientes.cell(row=row, column=5).fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
                    else:
                        self.aba_analitico_clientes.cell(row=row, column=4).fill = PatternFill(start_color='FF6565', end_color='FF6565', fill_type='solid')
                        self.aba_analitico_clientes.cell(row=row, column=5).fill = PatternFill(start_color='FF6565', end_color='FF6565', fill_type='solid')
                    if flag_data:
                        self.aba_analitico_clientes.cell(row=row, column=6).fill = PatternFill(start_color='d8e4bc', end_color='d8e4bc', fill_type='solid')
                        self.aba_analitico_clientes.cell(row=row, column=7).fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
                        self.aba_analitico_clientes.cell(row=row, column=8).fill = PatternFill(start_color='d8e4bc', end_color='d8e4bc', fill_type='solid')
                    else:
                        self.aba_analitico_clientes.cell(row=row, column=6).fill = PatternFill(start_color='e6b8b7', end_color='e6b8b7', fill_type='solid')
                        self.aba_analitico_clientes.cell(row=row, column=7).fill = PatternFill(start_color='FF6565', end_color='FF6565', fill_type='solid')
                        self.aba_analitico_clientes.cell(row=row, column=8).fill = PatternFill(start_color='e6b8b7', end_color='e6b8b7', fill_type='solid')

                    if flag_inv_pec:
                        self.aba_analitico_clientes.cell(row=row, column=10).fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
                    else:
                        self.aba_analitico_clientes.cell(row=row, column=10).fill = PatternFill(start_color='FF6565', end_color='FF6565', fill_type='solid')
                    if flag_inv_atend:
                        self.aba_analitico_clientes.cell(row=row, column=11).fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
                    else:
                        self.aba_analitico_clientes.cell(row=row, column=11).fill = PatternFill(start_color='FF6565', end_color='FF6565', fill_type='solid')
                    
                    if conteudo[4] == "-":
                        for col in (4,5,8,10,12):
                            self.aba_analitico_clientes.cell(row=row, column=col).fill = PatternFill(start_color='bfbfbf', end_color='bfbfbf', fill_type='solid')

                    if conteudo[9] == "-":
                        self.aba_analitico_clientes.cell(row=row, column=10).fill = PatternFill(start_color='bfbfbf', end_color='bfbfbf', fill_type='solid')
        
        # Definir tamanho das colunas
        column_width = 20
        for column_letter in ('A', 'B', 'F'):
            max_length = 0

            adjusted_width = (max_length + 2) * 1.2
            self.aba_analitico_clientes.column_dimensions[column_letter].width = adjusted_width if adjusted_width > column_width else column_width

        for colunas in ('C', 'L'):
            self.aba_analitico_clientes.column_dimensions[colunas].width = 60
        
        for colunas in ('G', 'H', 'I', 'J', 'K', 'D', 'E'):
            self.aba_analitico_clientes.column_dimensions[colunas].width = 25

        

        finish(self.aba_analitico_clientes)
        dateStyle(self.aba_analitico_clientes, 6)
        dateStyle(self.aba_analitico_clientes, 7)

    def sistematizar_analitico_atend(self):
       

        COLUNAS_ANALITICO_ATEND = ["UF", "MUNICIPIO", "UNIDADE", "N° LOTE", "DATA DO ULTIMO L. GERADO", "COMPETENCIA", "FICHAS VÁLIDAS", "FICHAS INVÁLIDAS"]

        header(self.aba_detalhado_atend, get_column_letter(len(COLUNAS_ANALITICO_ATEND)))

        for col, valor in enumerate(COLUNAS_ANALITICO_ATEND):
            celula = self.aba_detalhado_atend.cell(row=5, column=col+1, value=valor)
            celula.alignment = Alignment(
                horizontal='center', vertical='center')
            celula.font = Font(name='Calibri', size=11,
                               bold=True, color="000000")
            celula.fill = PatternFill(
                start_color="c0c0c0", end_color="c0c0c0", fill_type="solid")
            
        start_row = 5
        row = 5

        for municipio_atend, unidades_atend in self.lotes["ATEND"].items():            
            for unidade_atend, lotes_atend in unidades_atend.items():         
                for lote_atend in lotes_atend:
                    row+=1
                    self.aba_detalhado_atend.append(lote_atend.getDadosAtend())
                    try:
                        if int(lote_atend._invalid) > 0:
                            self.aba_detalhado_atend.cell(row=row, column=8).fill = PatternFill(start_color='FF6565', end_color='FF6565', fill_type='solid')
                    except ValueError:
                        pass
                   
        # Definir tamanho das colunas
        column_width = 20
        for column_letter in ('A', 'B', 'D', 'F', 'G', 'H'):
            max_length = 0

            adjusted_width = (max_length + 2) * 1.2
            self.aba_detalhado_atend.column_dimensions[column_letter].width = adjusted_width if adjusted_width > column_width else column_width

        for colunas in ('E'):
            self.aba_detalhado_atend.column_dimensions[colunas].width = 28

        for colunas in ('C'):
            self.aba_detalhado_atend.column_dimensions[colunas].width = 50
        
        finish(self.aba_detalhado_atend)  

    def sistematizar_analitico_pec(self):

        COLUNAS_ANALITICO_PEC = ["UF", "MUNICIPIO", "UNIDADE EXTRACAO", "UNIDADE", "N° LOTE", "ULTIMA IMPORTACAO", "HORA DA IMPORTACAO", "COMPETENCIA", "ENVIADO POR", "FICHAS VÁLIDAS", "FICHAS INVÁLIDAS", "TIPO DE IMPORTAÇÃO"]

        header(self.aba_detalhado_pec, get_column_letter(len(COLUNAS_ANALITICO_PEC)))

        for col, valor in enumerate(COLUNAS_ANALITICO_PEC):
            celula = self.aba_detalhado_pec.cell(row=5, column=col+1, value=valor)
            celula.alignment = Alignment(
                horizontal='center', vertical='center')
            celula.font = Font(name='Calibri', size=11,
                               bold=True, color="000000")
            celula.fill = PatternFill(
                start_color="c0c0c0", end_color="c0c0c0", fill_type="solid")
            
        row = 5

        for municipio_pec, unidades_pec in self.lotes["PEC"].items():
            for unidade_pec, lotes_pec in unidades_pec.items():
                flag_lote_incidencia = False
                for lote_pec in lotes_pec:
                    row+=1
                    flag_lote = False
                    flag_data = False
                    flag_importacao_invalida = False
                    flag_invalid = False
                    for municipio_atend, unidades_atend in self.lotes["ATEND"].items():
                        if municipio_pec == municipio_atend:
                            for unidade_atend, lotes_atend in unidades_atend.items():
                                if unidade_pec == unidade_atend:
                                    for lote_atend in lotes_atend:
                                        if lote_atend._nu_lote == lote_pec._nu_lote:
                                            flag_lote = True
                                            flag_lote_incidencia = True
                                            if lote_atend._data != lote_pec._data:
                                                flag_data = True
                                            break
                    try:                        
                        if int(lote_pec._invalid) > 0:
                            flag_invalid = True
                    except ValueError:
                        pass

                    self.aba_detalhado_pec.append(lote_pec.getDadosPec())

                    if flag_lote:
                        for col, valor in enumerate(COLUNAS_ANALITICO_PEC):
                            self.aba_detalhado_pec.cell(row=row, column=col+1).fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
                    if flag_data:
                        self.aba_detalhado_pec.cell(row=row, column=6).fill = PatternFill(start_color='FF6565', end_color='FF6565', fill_type='solid')
                    
                    if flag_invalid:
                            self.aba_detalhado_pec.cell(row=row, column=11).fill = PatternFill(start_color='FF6565', end_color='FF6565', fill_type='solid')

                if not flag_lote_incidencia:
                    if row <= 6:
                        start_rowFill = row
                        end_rowFill = row+(len(lotes_pec)-1)
                    else:
                        start_rowFill = row - (len(lotes_pec)-1)
                        end_rowFill = row + 1
                    for row_fill in range(start_rowFill, end_rowFill):
                        for col in (3, 4):
                            self.aba_detalhado_pec.cell(row=row_fill, column=col).fill = PatternFill(start_color='e6b8b7', end_color='e6b8b7', fill_type='solid')           

        # Definir tamanho das colunas
        column_width = 20
        for column_letter in ('A', 'B'):
            max_length = 0

            adjusted_width = (max_length + 2) * 1.2
            self.aba_detalhado_pec.column_dimensions[column_letter].width = adjusted_width if adjusted_width > column_width else column_width

        for colunas in ('C', 'D'):
            self.aba_detalhado_pec.column_dimensions[colunas].width = 50
        
        for colunas in ('E', 'F', 'G', 'H', 'I', 'J', 'K', 'L'):
            self.aba_detalhado_pec.column_dimensions[colunas].width = 25

        finish(self.aba_detalhado_pec)

if __name__ == '__main__':
    Sistematizar()